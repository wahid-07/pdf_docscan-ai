from fastapi import Depends, APIRouter, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse, Response, StreamingResponse
from pydantic import BaseModel
from auth.auth import get_current_user, get_current_admin
from services.pdf_handler import (
    get_all_pages_as_images,
    validate_pdf
)
from services.ai_extractor import extract_page_content
from services.db_handler import (
    create_pdf_master,
    update_pdf_status,
    insert_temp,
    move_temp_to_final,
    get_all_records,
    get_all_masters,
    delete_pdf_by_id,
    search_records,
    create_rejected_upload,
    get_rejected_uploads,
    calculate_file_sha256,
    find_duplicate_pdf,
    get_pdf_status,
)
from services.audit_logger import log_event
import asyncio
import re
import shutil
import logging
import os
import io
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()
logger = logging.getLogger("pdf_extractor.upload")


def _get_max_upload_size_mb() -> int:
    try:
        return int(os.getenv("MAX_UPLOAD_SIZE_MB", "500"))
    except ValueError:
        return 500


def _get_chunk_size_bytes() -> int:
    try:
        return int(os.getenv("CHUNK_SIZE_MB", "2")) * 1024 * 1024
    except ValueError:
        return 2 * 1024 * 1024


def _get_ai_concurrency() -> int:
    try:
        return max(1, int(os.getenv("AI_MAX_CONCURRENCY", "3")))
    except ValueError:
        return 3


def _is_valid_pdf_file(path: str) -> bool:
    try:
        with open(path, "rb") as handle:
            header = handle.read(8)
        return header.startswith(b"%PDF")
    except Exception:
        return False


async def _stream_upload_to_temp(file: UploadFile, temp_path: str, max_upload_size_mb: int, chunk_size: int) -> int:
    max_bytes = max_upload_size_mb * 1024 * 1024
    file_size = 0

    try:
        with open(temp_path, "wb") as buffer:
            while chunk := await file.read(chunk_size):
                file_size += len(chunk)
                if file_size > max_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail=f"File exceeds maximum size of {max_upload_size_mb} MB"
                    )
                buffer.write(chunk)
    except HTTPException:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise

    return file_size


async def _process_pdf_in_background(pdf_id: int, pdf_path: str, file_name: str, user_id: int | None, file_hash: str | None = None):
    try:
        update_pdf_status(pdf_id, "processing", progress=5, message="Preparing PDF")
        images = get_all_pages_as_images(pdf_path)
        total_pages = len(images)
        concurrency = _get_ai_concurrency()
        semaphore = asyncio.Semaphore(concurrency)

        async def extract_one(image, page_num):
            async with semaphore:
                return await asyncio.to_thread(extract_page_content, image, page_num)

        results = []
        for index, image in enumerate(images):
            progress = int(((index + 1) / total_pages) * 90) if total_pages else 90
            update_pdf_status(pdf_id, "processing", progress=progress, message=f"Extracting page {index + 1}")
            results.append(await extract_one(image, index + 1))

        for page_data in results:
            insert_temp(pdf_id, page_data)

        move_temp_to_final(pdf_id, file_name, user_id=user_id)
        update_pdf_status(pdf_id, "completed", progress=100, message="Finished")
    except Exception as exc:
        logging.exception("Background PDF processing failed for %s", file_name)
        update_pdf_status(pdf_id, "failed", progress=0, message="Processing failed")
        raise exc


# ── SINGLE UPLOAD ──
@router.post("/upload")
async def upload_pdf(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    background_tasks: BackgroundTasks = None,
):
    user_id = current_user.get("user_id")

    if not file.filename.endswith(".pdf"):
        logger.warning("Rejected non-PDF upload: %s", file.filename)
        raise HTTPException(status_code=400, detail="Only PDF files allowed")

    temp_path = f"temp_{file.filename}"
    max_upload_size_mb = _get_max_upload_size_mb()
    chunk_size = _get_chunk_size_bytes()
    file_size = await _stream_upload_to_temp(file, temp_path, max_upload_size_mb, chunk_size)

    try:
        validation = validate_pdf(temp_path)

        if not validation["valid"] or not _is_valid_pdf_file(temp_path):
            logger.warning("Upload rejected for %s: %s", file.filename, validation.get("reason"))

            reject_folder = os.path.join("uploads", "rejected", validation["folder"])
            os.makedirs(reject_folder, exist_ok=True)
            reject_path = os.path.join(reject_folder, file.filename)
            shutil.move(temp_path, reject_path)

            create_rejected_upload(
                file_name=file.filename,
                file_path=reject_path,
                reason=validation["reason"],
                file_size=file_size,
                user_id=user_id
            )

            return JSONResponse(
                status_code=400,
                content={
                    "status": "rejected",
                    "reason": validation["reason"],
                    "message": validation["message"]
                }
            )

        file_hash = calculate_file_sha256(temp_path)
        duplicate = find_duplicate_pdf(file_hash, user_id=user_id)
        if duplicate:
            logger.info("Duplicate upload rejected for %s", file.filename)
            log_event("duplicate_upload", user_id=user_id, details=file.filename)
            raise HTTPException(status_code=409, detail="Duplicate PDF already exists")

        accepted_folder = os.path.join("uploads", "accepted")
        os.makedirs(accepted_folder, exist_ok=True)
        accepted_path = os.path.join(accepted_folder, file.filename)
        shutil.copy(temp_path, accepted_path)

        with open(temp_path, "rb") as source:
            file_bytes = source.read()

        images = get_all_pages_as_images(accepted_path)
        total_pages = len(images)
        pdf_id = create_pdf_master(
            file.filename,
            total_pages,
            file_data=file_bytes,
            file_size=file_size,
            user_id=user_id,
            file_hash=file_hash,
        )

        if background_tasks is not None:
            background_tasks.add_task(
                _process_pdf_in_background,
                pdf_id,
                accepted_path,
                file.filename,
                user_id,
                file_hash,
            )

        logger.info("Upload accepted for %s with id %s", file.filename, pdf_id)
        log_event("upload_success", user_id=user_id, details=f"pdf_id={pdf_id} file={file.filename}")
        return JSONResponse({
            "status": "success",
            "processing": "background",
            "pdf_id": pdf_id,
            "file": file.filename,
            "total_pages": total_pages,
            "message": "PDF accepted and processing started in background"
        })

    except HTTPException:
        raise

    except Exception as e:
        logger.exception("Upload failed")

        raise HTTPException(
            status_code=500,
            detail="Internal Server Error"
        )

    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

# ── BULK UPLOAD ──
@router.post("/bulk-upload")
async def bulk_upload(
    files: list[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user),
    background_tasks: BackgroundTasks = None,
):
    user_id = current_user.get("user_id")
    results = []

    for file in files:
        if not file.filename.endswith(".pdf"):
            results.append({"file": file.filename, "status": "skipped", "reason": "Not a PDF"})
            continue

        temp_path = f"temp_{file.filename}"
        max_upload_size_mb = _get_max_upload_size_mb()
        chunk_size = _get_chunk_size_bytes()
        file_size = await _stream_upload_to_temp(file, temp_path, max_upload_size_mb, chunk_size)

        try:
            validation = validate_pdf(temp_path)

            if not validation["valid"] or not _is_valid_pdf_file(temp_path):
                reject_folder = os.path.join("uploads", "rejected", validation["folder"])
                os.makedirs(reject_folder, exist_ok=True)
                reject_path = os.path.join(reject_folder, file.filename)
                shutil.move(temp_path, reject_path)

                create_rejected_upload(
                    file_name=file.filename,
                    file_path=reject_path,
                    reason=validation["reason"],
                    file_size=file_size,
                    user_id=user_id
                )

                results.append({
                    "file": file.filename,
                    "status": "rejected",
                    "reason": validation["reason"]
                })
                continue

            file_hash = calculate_file_sha256(temp_path)
            duplicate = find_duplicate_pdf(file_hash, user_id=user_id)
            if duplicate:
                logger.info("Duplicate upload rejected for %s in bulk upload", file.filename)
                log_event("duplicate_upload", user_id=user_id, details=file.filename)
                results.append({
                    "file": file.filename,
                    "status": "rejected",
                    "reason": "Duplicate PDF already exists"
                })
                continue

            accepted_folder = os.path.join("uploads", "accepted")
            os.makedirs(accepted_folder, exist_ok=True)
            accepted_path = os.path.join(accepted_folder, file.filename)
            shutil.copy(temp_path, accepted_path)

            with open(temp_path, "rb") as source:
                file_bytes = source.read()

            images = get_all_pages_as_images(accepted_path)
            total_pages = len(images)
            pdf_id = create_pdf_master(
                file.filename,
                total_pages,
                file_data=file_bytes,
                file_size=file_size,
                user_id=user_id,
                file_hash=file_hash,
            )

            if background_tasks is not None:
                background_tasks.add_task(
                    _process_pdf_in_background,
                    pdf_id,
                    accepted_path,
                    file.filename,
                    user_id,
                    file_hash,
                )

            results.append({
                "file": file.filename,
                "status": "success",
                "processing": "background",
                "pdf_id": pdf_id,
                "pages": total_pages
            })

        except HTTPException as e:
            results.append({
                "file": file.filename,
                "status": "failed",
                "reason": e.detail
            })

        except Exception as e:
            logger.exception("Bulk upload failed")

            results.append({
                "file": file.filename,
                "status": "failed",
                "reason": "Internal Server Error"
            })
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    return JSONResponse({"results": results, "total": len(results)})


# ── GET PDF ──
@router.get("/pdf/{pdf_id}")
def get_pdf(
    pdf_id: int,
    current_user: dict = Depends(get_current_user)
):
    from database.connection import SessionLocal
    from models.table_model import PDFMaster

    user_id = current_user.get("user_id")
    role = current_user.get("role")

    db = SessionLocal()
    try:
        record = db.query(PDFMaster).filter(PDFMaster.id == pdf_id).first()

        if not record or not record.file_data:
            raise HTTPException(status_code=404, detail="PDF not found")

        # Ownership check
        if role != "admin" and record.user_id != user_id:
            raise HTTPException(status_code=403, detail="you can not access this pdf")

        return Response(
            content=bytes(record.file_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={record.file_name}"}
        )
    finally:
        db.close()


# ── DELETE PDF ──
@router.delete("/pdf/{pdf_id}")
def delete_pdf(
    pdf_id: int,
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user.get("user_id")
    role = current_user.get("role")

    result = delete_pdf_by_id(pdf_id, user_id=user_id, role=role)

    if result is None:
        raise HTTPException(status_code=403, detail="you can not delete this pdf")
    if result is False:
        raise HTTPException(status_code=404, detail="PDF not found")

    return {"status": "deleted", "pdf_id": pdf_id}


# ── GET ALL RECORDS ──
@router.get("/upload/status/{pdf_id}")
def get_upload_status(
    pdf_id: int,
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user.get("user_id")
    role = current_user.get("role")
    status = get_pdf_status(pdf_id, user_id=user_id, role=role)
    if not status:
        raise HTTPException(status_code=404, detail="PDF not found")
    return status


@router.get("/admin/overview")
def admin_overview(current_user: dict = Depends(get_current_admin)):
    from database.connection import SessionLocal
    from models.table_model import PDFMaster, User

    db = SessionLocal()
    try:
        total_users = db.query(User).count()
        total_pdfs = db.query(PDFMaster).count()
        completed = db.query(PDFMaster).filter(PDFMaster.status == "completed").count()
        failed = db.query(PDFMaster).filter(PDFMaster.status == "failed").count()
        return {
            "total_users": total_users,
            "total_pdfs": total_pdfs,
            "completed": completed,
            "failed": failed,
        }
    finally:
        db.close()


@router.get("/admin/users")
def admin_list_users(current_user: dict = Depends(get_current_admin)):
    from database.connection import SessionLocal
    from models.table_model import User

    db = SessionLocal()
    try:
        users = db.query(User).order_by(User.id).all()
        return {
            "total": len(users),
            "users": [
                {
                    "id": u.id,
                    "full_name": u.full_name,
                    "email": u.email,
                    "role": u.role,
                    "is_active": u.is_active,
                    "account_locked": u.account_locked,
                    "created_at": str(u.created_at) if u.created_at else None,
                    "last_login": str(u.last_login) if u.last_login else None,
                }
                for u in users
            ],
        }
    finally:
        db.close()


class RoleUpdateRequest(BaseModel):
    role: str


@router.put("/admin/users/{user_id}/role")
def admin_update_user_role(
    user_id: int,
    data: RoleUpdateRequest,
    current_user: dict = Depends(get_current_admin),
):
    from database.connection import SessionLocal
    from models.table_model import User

    if data.role not in ("user", "admin"):
        raise HTTPException(status_code=400, detail="Role must be 'user' or 'admin'.")

    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == user_id).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found.")

        # Prevent an admin from demoting themselves and locking everyone out
        if target.id == current_user.get("user_id") and data.role != "admin":
            raise HTTPException(status_code=400, detail="You cannot remove your own admin role.")

        # Prevent removing the last remaining admin
        if target.role == "admin" and data.role == "user":
            remaining_admins = db.query(User).filter(User.role == "admin", User.id != target.id).count()
            if remaining_admins == 0:
                raise HTTPException(status_code=400, detail="Cannot demote the last remaining admin.")

        target.role = data.role
        db.commit()
        log_event("role_change", user_id=current_user.get("user_id"), details=f"target_user={target.id} new_role={data.role}")
        return {"success": True, "message": f"{target.email} is now {data.role}.", "user_id": target.id, "role": target.role}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


class StatusUpdateRequest(BaseModel):
    is_active: bool


@router.put("/admin/users/{user_id}/status")
def admin_update_user_status(
    user_id: int,
    data: StatusUpdateRequest,
    current_user: dict = Depends(get_current_admin),
):
    from database.connection import SessionLocal
    from models.table_model import User

    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == user_id).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found.")

        if target.id == current_user.get("user_id") and not data.is_active:
            raise HTTPException(status_code=400, detail="You cannot deactivate your own account.")

        if target.role == "admin" and not data.is_active:
            remaining_active_admins = db.query(User).filter(
                User.role == "admin", User.is_active == True, User.id != target.id  # noqa: E712
            ).count()
            if remaining_active_admins == 0:
                raise HTTPException(status_code=400, detail="Cannot deactivate the last active admin.")

        target.is_active = data.is_active
        db.commit()
        action = "activated" if data.is_active else "deactivated"
        log_event("user_status_change", user_id=current_user.get("user_id"), details=f"target_user={target.id} action={action}")
        return {"success": True, "message": f"{target.email} {action}.", "user_id": target.id, "is_active": target.is_active}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


@router.delete("/admin/users/{user_id}")
def admin_delete_user(
    user_id: int,
    current_user: dict = Depends(get_current_admin),
):
    """
    User delete karta hai. Unke PDFs/records/rejected-uploads delete nahi hote
    (data loss se bachne ke liye) — sirf un rows ka user_id NULL kar diya jata hai,
    taaki wo "orphaned but preserved" rahein, foreign key error na aaye.
    """
    from database.connection import SessionLocal
    from models.table_model import User, PDFMaster, ExtractedData, RejectedUpload

    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == user_id).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found.")

        if target.id == current_user.get("user_id"):
            raise HTTPException(status_code=400, detail="You cannot delete your own account.")

        if target.role == "admin":
            remaining_admins = db.query(User).filter(User.role == "admin", User.id != target.id).count()
            if remaining_admins == 0:
                raise HTTPException(status_code=400, detail="Cannot delete the last remaining admin.")

        db.query(PDFMaster).filter(PDFMaster.user_id == target.id).update({"user_id": None})
        db.query(ExtractedData).filter(ExtractedData.user_id == target.id).update({"user_id": None})
        db.query(RejectedUpload).filter(RejectedUpload.user_id == target.id).update({"user_id": None})

        target_email = target.email
        db.delete(target)
        db.commit()
        log_event("user_deleted", user_id=current_user.get("user_id"), details=f"deleted_email={target_email}")
        return {"success": True, "message": f"{target_email} deleted."}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


@router.get("/admin/audit-logs")
def admin_audit_logs(current_user: dict = Depends(get_current_admin)):
    """
    audit.log lines format: "2026-07-14 10:22:01 - event=login user_id=3 details=email=a@b.com"
    Isko structured fields (timestamp, event, user_id, details) mein todta hai,
    taaki frontend ek proper table bana sake, raw text blob nahi.
    """
    log_path = Path("logs/audit.log")
    if not log_path.exists():
        return {"total": 0, "logs": []}
    lines = log_path.read_text(encoding="utf-8").splitlines()[-200:]
    logs = []
    for line in lines:
        if not line.strip():
            continue
        parts = line.split(" - ", 1)
        timestamp = parts[0] if parts else ""
        message = parts[1] if len(parts) > 1 else line

        event = "unknown"
        log_user_id = "system"
        details = ""

        event_match = re.search(r"event=(\S+)", message)
        if event_match:
            event = event_match.group(1)

        user_match = re.search(r"user_id=(\S+)", message)
        if user_match:
            log_user_id = user_match.group(1)

        details_match = re.search(r"details=(.*)$", message)
        if details_match:
            details = details_match.group(1)

        logs.append({
            "timestamp": timestamp,
            "event": event,
            "user_id": log_user_id,
            "details": details,
            "raw": message,
        })
    return {"total": len(logs), "logs": logs}


@router.get("/records")
def get_records(
    search: str = None,
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user.get("user_id")
    role = current_user.get("role")

    if search:
        records = search_records(search, user_id=user_id, role=role)
    else:
        records = get_all_records(user_id=user_id, role=role)

    return {
        "total": len(records),
        "records": [
            {
                "id": r.id,
                "pdf_id": r.pdf_id,
                "file_name": r.file_name,
                "page_number": r.page_number,
                "content_type": r.content_type,
                "data": r.data,
                "uploaded_at": str(r.uploaded_at)
            }
            for r in records
        ]
    }


# ── GET ALL MASTERS ──
@router.get("/masters")
def get_masters(
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user.get("user_id")
    role = current_user.get("role")

    masters = get_all_masters(user_id=user_id, role=role)

    return {
        "total": len(masters),
        "pdfs": [
            {
                "id": m.id,
                "file_name": m.file_name,
                "total_pages": m.total_pages,
                "status": m.status,
                "file_size": m.file_size,
                "uploaded_at": str(m.uploaded_at)
            }
            for m in masters
        ]
    }


# ── GET REJECTED PDFS ──
@router.get("/rejected")
def get_rejected(
    current_user: dict = Depends(get_current_user)
):
    from database.connection import SessionLocal
    from models.table_model import User

    user_id = current_user.get("user_id")
    role = current_user.get("role")

    rejected = get_rejected_uploads(user_id=user_id, role=role)

    db = SessionLocal()
    try:
        owner_ids = {r.user_id for r in rejected if r.user_id}
        owners = {u.id: u.email for u in db.query(User).filter(User.id.in_(owner_ids)).all()} if owner_ids else {}
    finally:
        db.close()

    return {
        "total": len(rejected),
        "rejected": [
            {
                "id": r.id,
                "file_name": r.file_name,
                "file_path": r.file_path,
                "reason": r.reason,
                "file_size": r.file_size,
                "uploaded_at": str(r.uploaded_at),
                "owner_email": owners.get(r.user_id, "—"),
            }
            for r in rejected
        ]
    }


@router.get("/rejected/{rejected_id}/file")
def get_rejected_file(
    rejected_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Rejected PDF ko disk se serve karta hai (preview ke liye).
    RBAC: owner ya admin hi access kar sakta hai.
    """
    from database.connection import SessionLocal
    from models.table_model import RejectedUpload

    user_id = current_user.get("user_id")
    role = current_user.get("role")

    db = SessionLocal()
    try:
        record = db.query(RejectedUpload).filter(RejectedUpload.id == rejected_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Rejected file not found")

        if role != "admin" and record.user_id != user_id:
            raise HTTPException(status_code=403, detail="You cannot access this file")

        if not os.path.exists(record.file_path):
            raise HTTPException(status_code=404, detail="File no longer exists on disk")

        with open(record.file_path, "rb") as f:
            content = f.read()

        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={record.file_name}"}
        )
    finally:
        db.close()


# ── DASHBOARD STATS ──
@router.get("/stats")
def get_stats(
    current_user: dict = Depends(get_current_user)
):
    from database.connection import SessionLocal
    from models.table_model import PDFMaster, ExtractedData
    from sqlalchemy import func

    user_id = current_user.get("user_id")
    role = current_user.get("role")

    db = SessionLocal()
    try:
        pdf_query = db.query(PDFMaster)
        rec_query = db.query(ExtractedData)

        # User apna data dekhe, admin sab dekhe
        if role != "admin":
            pdf_query = pdf_query.filter(PDFMaster.user_id == user_id)
            rec_query = rec_query.filter(ExtractedData.user_id == user_id)

        total_pdfs    = pdf_query.count()
        completed     = pdf_query.filter(PDFMaster.status == "completed").count()
        failed        = pdf_query.filter(PDFMaster.status == "failed").count()
        total_pages   = db.query(func.sum(PDFMaster.total_pages)).scalar() or 0
        total_records = rec_query.count()

        type_counts = rec_query.with_entities(
            ExtractedData.content_type,
            func.count(ExtractedData.id)
        ).group_by(ExtractedData.content_type).all()

        recent = pdf_query.order_by(PDFMaster.uploaded_at.desc()).limit(5).all()

        return {
            "total_pdfs": total_pdfs,
            "completed": completed,
            "failed": failed,
            "total_pages": total_pages,
            "total_records": total_records,
            "content_types": {t: c for t, c in type_counts},
            "recent_uploads": [
                {
                    "id": r.id,
                    "file_name": r.file_name,
                    "pages": r.total_pages,
                    "status": r.status,
                    "uploaded_at": str(r.uploaded_at)
                }
                for r in recent
            ]
        }
    finally:
        db.close()


# ── EXCEL EXPORT ──
@router.get("/export/excel")
def export_excel(
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user.get("user_id")
    role = current_user.get("role")

    records = get_all_records(user_id=user_id, role=role)
    if not records:
        raise HTTPException(status_code=404, detail="Koi data nahi hai")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "File Name", "PDF Link", "Page No", "Content Type", "Text Preview", "Tables", "Images", "Uploaded At"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, r in enumerate(records, 2):
        text_preview = ""
        tables_count = 0
        images_count = 0

        if r.data:
            text_preview = str(r.data.get("text", ""))[:200]
            tables_count = len(r.data.get("tables", []))
            images_count = len(r.data.get("images", []))

        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.file_name)

        pdf_url = f"http://127.0.0.1:8000/api/pdf/{r.pdf_id}"
        link_cell = ws.cell(row=row_idx, column=3, value="Open PDF")
        link_cell.hyperlink = pdf_url
        link_cell.style = "Hyperlink"

        ws.cell(row=row_idx, column=4, value=r.page_number)
        ws.cell(row=row_idx, column=5, value=r.content_type)
        ws.cell(row=row_idx, column=6, value=text_preview)
        ws.cell(row=row_idx, column=7, value=tables_count)
        ws.cell(row=row_idx, column=8, value=images_count)
        ws.cell(row=row_idx, column=9, value=str(r.uploaded_at))

        if row_idx % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F0F4FF", end_color="F0F4FF", fill_type="solid"
                )

    widths = [6, 30, 20, 8, 15, 50, 14, 14, 22]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=docscan_export.xlsx"}
    )